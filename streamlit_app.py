import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# قائمة الفروع المسموح بها
allowed_branches = {
    "NBIQIQBA830",
    "NBIQIQBA856",
    "NBIQIQBA859",
    "NBIQIQBA005",
    "NBIQIQBA860",
    "NBIQIQBA862",
    "NBIQIQBA849",
    "NBIQIQBA865",
    "NBIQIQBA844",
    "NBIQIQBA848",
    "NBIQIQBA850"
}

def extract_branch_from_iban(iban):
    try:
        return iban[8:11]
    except:
        return ""

def correct_bic(row):
    try:
        branch = extract_branch_from_iban(row['IBAN'])
        bic = f"NBIQIQBA{branch}"
        return bic if bic in allowed_branches else row['Receiver BIC']
    except:
        return row['Receiver BIC']

def process_excel(file):
    df = pd.read_excel(file)

    if 'IBAN' not in df.columns or 'Receiver BIC' not in df.columns:
        st.error("يجب أن يحتوي الملف على عمودين: IBAN و Receiver BIC")
        return None, 0

    original_bic = df['Receiver BIC'].copy()
    df['Receiver BIC'] = df.apply(correct_bic, axis=1)
    modified_rows = (original_bic != df['Receiver BIC']).sum()

    # حفظ الملف في BytesIO
    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)

    # فتحه بـ openpyxl للتنسيق
    wb = load_workbook(output)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    bic_col_idx = None
    for i, col in enumerate(ws[1], start=1):
        if col.value == 'Receiver BIC':
            bic_col_idx = i
            break

    if bic_col_idx:
        for row_idx, (original, corrected) in enumerate(zip(original_bic, df['Receiver BIC']), start=2):
            if original != corrected:
                ws.cell(row=row_idx, column=bic_col_idx).fill = yellow_fill

    # ضبط عرض الأعمدة
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        column_letter = get_column_letter(column)
        for cell in column_cells:
            try:
                cell_value = str(cell.value)
                if cell_value:
                    max_length = max(max_length, len(cell_value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    # حفظ إلى BytesIO جديد
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return final_output, modified_rows

# واجهة Streamlit
st.set_page_config(page_title="تصحيح BIC", layout="centered")
st.title("🧮 تصحيح رمز الفرع في Receiver BIC")
st.write("يرجى رفع ملف Excel يحتوي على عمودي `IBAN` و `Receiver BIC` لتصحيحه.")

uploaded_file = st.file_uploader("📁 ارفع ملف Excel", type=["xlsx", "xls"])

if uploaded_file:
    if st.button("🔧 بدء التصحيح"):
        with st.spinner("جاري المعالجة..."):
            processed_file, modified_count = process_excel(uploaded_file)

        if processed_file:
            st.success(f"✅ تم تعديل {modified_count} صف/صفوف.")
            st.download_button(
                label="📥 تحميل الملف المعدل",
                data=processed_file,
                file_name="BIC_Corrected.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
